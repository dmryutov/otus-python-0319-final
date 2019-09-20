from itertools import zip_longest

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, Side


STYLE_CENTER = Alignment(horizontal='center')
STYLE_TITLE = Font(bold=True, size=14)
STYLE_BOLD = Font(bold=True)


class Excel(object):
    """
    Export project items to excel wrapper
    """
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    @staticmethod
    def emu_width(width):
        """
        Convert width in pixels to EMU

        Args:
            width (float): Width (in pixels)

        Returns:
            float: Width in EMU
        """
        return width * 9525  # Width in EMUs: pixel * 914400 / 96

    @staticmethod
    def set_border(cells):
        """
        Set borders to range of cells

        Args:
            cells: Range of cells
        """
        side = Side(border_style='thin', color='000000')
        border = Border(left=side, right=side, top=side, bottom=side)
        for row in cells:
            for cell in row:
                cell.border = border

    def export_stl(self, data, result):
        """
        Export time series decomposition results

        Args:
            data (list[float]): Input values
            result (dict): Time series decomposition results

        Returns:
            openpyxl.Workbook: Excel workbook
        """
        ws = self.wb.create_sheet('STL')

        # Styles
        titles = ('Observed', 'Trend', 'Seasonal', 'Residual')
        for i, letter in enumerate(('A', 'B', 'C', 'D')):
            cell = '{}1'.format(letter)
            ws.column_dimensions[letter].width = 12
            ws[cell] = titles[i]
            ws[cell].font = STYLE_BOLD
            ws[cell].alignment = STYLE_CENTER

        # Values
        data_length = len(data)
        row_data = zip(data, result['trend'], result['seasonal'], result['resid'])
        for row_idx, (x, trend, seasonal, resid) in enumerate(row_data, 2):
            ws['A{}'.format(row_idx)] = x
            ws['B{}'.format(row_idx)] = trend
            ws['C{}'.format(row_idx)] = seasonal
            ws['D{}'.format(row_idx)] = resid
        self.set_border(ws['A1:D{}'.format(data_length + 1)])

        # Charts
        for i, title in enumerate(titles):
            chart = LineChart()
            chart.title = title
            chart.style = 10
            chart.width = 20  # cm
            chart.height = 5  # cm
            chart.x_axis.tickLblPos = 'low'

            y_axis = Reference(ws, min_col=i + 1, min_row=1, max_row=data_length + 1)
            chart.add_data(y_axis, titles_from_data=True)

            ws.add_chart(chart, 'F{}'.format(i * 10 + 1))

        return self.wb

    def export_forecast(self, data, result, date_start, period_type):
        """
        Export time series forecasting results

        Args:
            data (list[float]): Input values
            result (dict): Time series forecasting results
            date_start (str): Time series first date
            period_type (str): Time series period type

        Returns:
            openpyxl.Workbook: Excel workbook
        """
        ws = self.wb.create_sheet('Forecast')

        # Styles
        titles = ('Date', 'Observed', 'Forecast', 'Conf. Lower', 'Conf. Upper')
        for i, letter in enumerate(('A', 'B', 'C', 'D', 'E')):
            cell = '{}1'.format(letter)
            ws.column_dimensions[letter].width = 12
            ws[cell] = titles[i]
            ws[cell].font = STYLE_BOLD
            ws[cell].alignment = STYLE_CENTER

        # Values
        data_length = len(result['forecast'])
        dates = pd.date_range(date_start, periods=data_length, freq=period_type)
        row_data = zip_longest(dates, data, result['forecast'], result['lower'], result['upper'])
        for row_idx, (dt, x, forecast, lower, upper) in enumerate(row_data, 2):
            ws['A{}'.format(row_idx)] = dt.date()
            ws['B{}'.format(row_idx)] = x
            ws['C{}'.format(row_idx)] = forecast
            ws['D{}'.format(row_idx)] = lower
            ws['E{}'.format(row_idx)] = upper
        self.set_border(ws['A1:E{}'.format(data_length + 1)])

        # Charts
        chart = LineChart()
        chart.title = 'Forecasting'
        chart.style = 10
        chart.width = 20  # cm
        chart.height = 10  # cm
        chart.x_axis.tickLblPos = 'low'

        x_axis = Reference(ws, min_col=1, min_row=2, max_row=data_length + 1)
        y_axis = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=data_length + 1)
        chart.add_data(y_axis, titles_from_data=True)
        chart.set_categories(x_axis)

        ws.add_chart(chart, 'G1')

        return self.wb
